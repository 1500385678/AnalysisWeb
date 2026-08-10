"""写 Coder sheet Row 77 · 批 7 R217 完成记录"""
from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook('/Users/aaron/Mac/项目夜间迭代-2026-08.xlsx', data_only=False)
ws = wb['Coder']
new_row = ws.max_row + 1
now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
start_time = '2026-08-10 23:40:00'

ws.cell(row=new_row, column=1, value=start_time)
ws.cell(row=new_row, column=2, value='analysisweb')
ws.cell(row=new_row, column=3, value='analysisweb-coder')
ws.cell(row=new_row, column=4, value='P0')
ws.cell(row=new_row, column=5, value='R217 server.py:9 ANALYSISWEB_HOME 默认值仍写 D:/Mac/... Windows 路径,build_db.py:48 已有 Mac 守卫,server.py 缺 → Mac mini 没设 env 启动 IMG_ROOT 错误路径所有图片 404;抄 build_db.py:48 模板加守卫:darwin + startswith("D:") → print 提示 + sys.exit(1);同步 AGENTS.md §15 批 7 变更记录')
ws.cell(row=new_row, column=6, value='ea68160 + fa2a14b + f538ff0 + 34506c0')
ws.cell(row=new_row, column=7, value='✅ (ceb094e / cd0ab04 / 6c431b3 / 6ea7855,git_data_push.py Git Data API)')
ws.cell(row=new_row, column=8, value='✅ (8cd1232a8b / 2354a4f542 / d441adef99 / b6332bc450 / f22f86dd1d,_push_fix_r217.py Contents API)')
ws.cell(row=new_row, column=9, value=now)
ws.cell(row=new_row, column=10, value="J10 审签通过(司令员'不要询问'),6 条护栏临时豁免,直接执行;cron 任务'选 1 条 P0 优先同P取第1' → analysisweb 8-10 23:25 批共 5 条意见(R217 P0 server.py Mac 守卫 / R218 P0 limit 解析 + LIMIT 注入 / R219 P0 _ai_image innerHTML 拼 / R220 P1 do_POST Content-Length 解析 / R221 P1 AGENTS.md 工作目录/图片根),按'P0 优先同P取第1'取 R217(本批首个 P0);R217 修复 3 处:1) server.py:11-15 加 Mac 守卫(6 行),抄 build_db.py:48 模板;2) AGENTS.md §15 加批 7 变更记录(11 行表);3) 新建 scripts/_push_fix_r217.py Gitee Contents API 推送工具(后续 Gitee 一文件一 commit 复用);验证:python3 server.py(无 env)→ exit=1 + 提示;ANALYSISWEB_HOME=/Users/aaron/.../Attack/03-Analysis python3 -c 'import server'→ IMG_ROOT 正确解析;wc -l server.py 811→817 + AGENTS.md 219→229;git push origin main 第1次+retry 1次 75s 超时(本机 github.com:443 网络层连不通,跟 R70 commanddashboard / R207 macweb 同症),fallback 走 git_data_push.py Git Data API + 自己写的 _push_fix_r217.py Gitee Contents API 双仓成功;8-9/8-10 Verifier R82/R83/R84/R85/R142/R143/R144/R145 此前已闭环(R17/R18/R19/R20/R21 是 8-7 批已闭环),R218/R219/R220/R221 待下一轮;sessionId=mvs_a7bc284e943041719c167b9b08d42295")

wb.save('/Users/aaron/Mac/项目夜间迭代-2026-08.xlsx')
print(f"✅ Coder sheet Row {new_row} 写入完成")
print(f"   时间={start_time}")
print(f"   完成时间={now}")
